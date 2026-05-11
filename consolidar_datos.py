import os
import pandas as pd
import json
from glob import glob
from datetime import datetime
import re

# --- CONFIGURACIÓN ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETA_RESULTADOS = os.path.join(BASE_DIR, "Resultados_scraping")
OUTPUT_JSON = os.path.join(BASE_DIR, "dashboard", "public", "datos_consolidados.json")
# Ruta al Excel oficial
RUTA_KILLERS = r"C:\Users\b_saravia\OneDrive - Farmacity\Documents\skillers\Killers evento.xlsx"

# Lista de farmacias soportadas
FARMACIAS = [
    ("Central_Oeste", "CentralOeste"),
    ("Farmacity",     "Farmacity"),
    ("Farmaonline",   "Farmaonline"),
]

print("--- INICIANDO CONSOLIDACIÓN ---")

archivos = glob(os.path.join(CARPETA_RESULTADOS, "*.xlsx"))
archivos = [f for f in archivos if "Comparacion_" not in os.path.basename(f) and not os.path.basename(f).startswith("~$")]

dfs = {}
for nombre_farmacia, patron in FARMACIAS:
    encontrados = sorted([f for f in archivos if patron in f], reverse=True)
    for f_path in encontrados:
        try:
            df = pd.read_excel(f_path)
            df['Farmacia'] = nombre_farmacia
            dfs[nombre_farmacia] = df
            print(f"Cargado {nombre_farmacia}: {os.path.basename(f_path)} ({len(df)} productos)")
            break # Si cargó bien, pasamos a la siguiente farmacia
        except Exception as e:
            print(f"Advertencia: El archivo {os.path.basename(f_path)} está corrupto o es inválido. Saltando... Error: {e}")
            continue

# --- CARGAR EXCEL OFICIAL ---
df_killers = pd.DataFrame()
if os.path.exists(RUTA_KILLERS):
    try:
        df_killers = pd.read_excel(RUTA_KILLERS, dtype={'EAN': str})
        df_killers['EAN'] = df_killers['EAN'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        print(f"Cargado Excel Oficial (Killers Evento): {len(df_killers)} productos Killers")
    except Exception as e:
        print(f"Error leyendo Killers Evento (Asegúrate que el archivo no esté abierto): {e}")
else:
    print(f"No se encontró el archivo en {RUTA_KILLERS}")

if not dfs:
    print("No se encontraron archivos de scraping. Abortando.")
    exit()

# Concatenar todo para procesar
fuentes = list(dfs.values())
df_total = pd.concat(fuentes, ignore_index=True)

# --- INDEXAR RESULTADOS DEL SCRAPING POR EAN/NOMBRE ---
scraping_index = {}

# Concatenar todos los resultados de scraping disponibles
fuentes = list(dfs.values())
df_total = pd.concat(fuentes, ignore_index=True)

for _, row in df_total.iterrows():
    # Usamos las nuevas columnas EAN_Original y Nombre_Original si existen, sino hacemos fallback a las normales
    ean_orig = str(row.get('EAN_Original', row.get('EAN', '')))
    nom_orig = str(row.get('Nombre_Original', row.get('Nombre', '')))
    
    if 'e' in ean_orig.lower() or '.' in ean_orig:
        try: ean_orig = format(float(ean_orig), '.0f')
        except: pass
    ean_val = re.sub(r'\.0$', '', ean_orig).strip().lower()
    nombre_val = nom_orig.lower().strip()
    
    # Usamos EAN original como clave principal, o Nombre original como secundaria
    claves = []
    if ean_val and ean_val != 'nan': claves.append(ean_val)
    if nombre_val: claves.append(nombre_val)
    
    for c in claves:
        if c not in scraping_index:
            scraping_index[c] = {"Precios": {}}
        
        farmacia = row['Farmacia']
        scraping_index[c]["Precios"][farmacia] = {
            "Precio_Final": row['Precio_Final'],
            "Precio_Lista": row['Precio_Lista'],
            "Descuento": row.get('Porcentaje_Oferta', 0),
            "Stock": row.get('Stock', 'Con Stock'),
            "Link": row.get('Link', '#')
        }

# --- CONSTRUIR LISTA FINAL BASADA EN EL EXCEL SAGRADO ---
lista_final = []
nombres_farmacias = [f[0] for f in FARMACIAS]

if not df_killers.empty:
    col_nombre_k = [c for c in df_killers.columns if 'Descripci' in c]
    col_nombre_k = col_nombre_k[0] if col_nombre_k else 'Nombre'
    
    for idx, row in df_killers.iterrows():
        ean_k = str(row.get('EAN', '')).strip()
        if 'e' in ean_k.lower() or '.' in ean_k:
            try: ean_k = format(float(ean_k), '.0f')
            except: pass
        ean_k = re.sub(r'\.0$', '', ean_k).strip()
        
        nombre_k = str(row.get(col_nombre_k, row.get('Nombre', ''))).strip()
        
        # Intentamos encontrar los precios en nuestro índice de scraping
        clave_ean = ean_k.lower() if ean_k and ean_k != 'nan' else None
        clave_nom = nombre_k.lower().strip()
        
        datos_scraping = None
        if clave_ean and clave_ean in scraping_index:
            datos_scraping = scraping_index[clave_ean]
        elif clave_nom and clave_nom in scraping_index:
            datos_scraping = scraping_index[clave_nom]
        
        precios_encontrados = datos_scraping["Precios"] if datos_scraping else {}
        
        producto_consolidado = {
            "Id": f"row_{idx}", # ID único por fila para evitar colapsos
            "EAN": ean_k,
            "Nombre": nombre_k,
            "Grupo": "Killers",
            "Es_Killer": True,
            "Precios": precios_encontrados
        }
        
        # --- Determinar ganador para esta fila ---
        precios_validos = {}
        for f_nom in nombres_farmacias:
            d_farm = precios_encontrados.get(f_nom, {})
            pf = d_farm.get("Precio_Final", float('inf'))
            stk = str(d_farm.get("Stock", "")).strip().lower()
            
            if stk == "sin stock" or str(pf).strip().lower() == "sin stock":
                precios_validos[f_nom] = float('inf')
            else:
                try:
                    pf = float(pf)
                    if pf > 0: precios_validos[f_nom] = pf
                    else: precios_validos[f_nom] = float('inf')
                except:
                    precios_validos[f_nom] = float('inf')

        f_con_p = {n: p for n, p in precios_validos.items() if p != float('inf')}
        
        if not f_con_p:
            producto_consolidado["Ganador"] = "Sin precio"
            producto_consolidado["MejorPrecio"] = 0
        elif len(f_con_p) == 1:
            u = list(f_con_p.keys())[0]
            producto_consolidado["Ganador"] = f"Exclusivo {u.replace('_', ' ')}"
            producto_consolidado["MejorPrecio"] = f_con_p[u]
        else:
            min_p = min(f_con_p.values())
            wins = [n for n, p in f_con_p.items() if p == min_p]
            producto_consolidado["Ganador"] = "Empate" if len(wins) > 1 else wins[0].replace("_", " ")
            producto_consolidado["MejorPrecio"] = min_p
            
            max_p = max(f_con_p.values())
            producto_consolidado["Diferencia_Porcentual"] = round(((max_p - min_p) / max_p) * 100, 1) if max_p > 0 else 0

        producto_consolidado["Disponible_En"] = len(f_con_p)
        lista_final.append(producto_consolidado)
else:
    print("ERROR: No hay lista base de Killers para procesar.")

# --- Guardar JSON para dashboard ---
os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
data_dashboard = {
    "ultima_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    "productos": lista_final
}
import math
def clean_nans(obj):
    if isinstance(obj, list):
        return [clean_nans(i) for i in obj]
    elif isinstance(obj, dict):
        return {k: clean_nans(v) for k, v in obj.items()}
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
    return obj

data_limpia = clean_nans(data_dashboard)

with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(data_limpia, f, ensure_ascii=False, indent=4)

print(f"Consolidación exitosa. JSON guardado en: {OUTPUT_JSON}")

# --- EXPORTAR A EXCEL (tabla comparativa) ---
filas_excel = []
for v in lista_final:
    precios = v.get("Precios", {})
    co = precios.get("Central_Oeste", {})
    fa = precios.get("Farmacity", {})
    fo = precios.get("Farmaonline", {})

    filas_excel.append({
        "EAN":                          v.get("EAN", "N/A"),
        "Nombre":                       v.get("Nombre", ""),
        "CO - Precio Final":            co.get("Precio_Final", ""),
        "CO - Precio Lista":            co.get("Precio_Lista", ""),
        "CO - Descuento %":             co.get("Descuento", ""),
        "CO - Stock":                   co.get("Stock", "No Encontrado"),
        "FA - Precio Final":            fa.get("Precio_Final", ""),
        "FA - Precio Lista":            fa.get("Precio_Lista", ""),
        "FA - Descuento %":             fa.get("Descuento", ""),
        "FA - Stock":                   fa.get("Stock", "No Encontrado"),
        "FO - Precio Final":            fo.get("Precio_Final", ""),
        "FO - Precio Lista":            fo.get("Precio_Lista", ""),
        "FO - Descuento %":             fo.get("Descuento", ""),
        "FO - Stock":                   fo.get("Stock", "No Encontrado"),
        "Ganador":                      v.get("Ganador", ""),
        "Mejor Precio":                 v.get("MejorPrecio", ""),
        "Diferencia %":                 v.get("Diferencia_Porcentual", 0),
        "Link CO":                      co.get("Link", ""),
        "Link FA":                      fa.get("Link", ""),
        "Link FO":                      fo.get("Link", ""),
    })

df_excel = pd.DataFrame(filas_excel)

ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(CARPETA_RESULTADOS, f"Comparacion_Farmacias_{ts}.xlsx")

with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    df_excel.to_excel(writer, index=False, sheet_name="Comparación")

    ws = writer.sheets["Comparación"]
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fill_header  = PatternFill("solid", fgColor="1E293B")
    fill_co      = PatternFill("solid", fgColor="DBEAFE")
    fill_fa      = PatternFill("solid", fgColor="FCE7F3")
    fill_fo      = PatternFill("solid", fgColor="FEF3C7")
    fill_winner  = PatternFill("solid", fgColor="DCFCE7")
    fill_co_win  = PatternFill("solid", fgColor="93C5FD")
    fill_fa_win  = PatternFill("solid", fgColor="F9A8D4")
    fill_fo_win  = PatternFill("solid", fgColor="FCD34D")
    font_header  = Font(bold=True, color="F8FAFC")
    font_link    = Font(color="2563EB", underline="single")
    thin_border  = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # 1-indexed
    col_co_pf = 3;  col_co_pl = 4;  col_co_dc = 5; col_co_st = 6
    col_fa_pf = 7;  col_fa_pl = 8;  col_fa_dc = 9; col_fa_st = 10
    col_fo_pf = 11; col_fo_pl = 12; col_fo_dc = 13; col_fo_st = 14
    col_ganador = 15; col_mejor = 16; col_dif = 17
    col_lk_co = 18; col_lk_fa = 19; col_lk_fo = 20

    cols_co = [col_co_pf, col_co_pl, col_co_dc, col_co_st, col_lk_co]
    cols_fa = [col_fa_pf, col_fa_pl, col_fa_dc, col_fa_st, col_lk_fa]
    cols_fo = [col_fo_pf, col_fo_pl, col_fo_dc, col_fo_st, col_lk_fo]

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        ganador_val = ws.cell(row=row_idx, column=col_ganador).value or ""
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            
            if col_idx in cols_co: cell.fill = fill_co
            elif col_idx in cols_fa: cell.fill = fill_fa
            elif col_idx in cols_fo: cell.fill = fill_fo
            elif col_idx == col_ganador:
                if "Central" in ganador_val: cell.fill = fill_co_win
                elif "Farmacity" in ganador_val: cell.fill = fill_fa_win
                elif "Farmaonline" in ganador_val: cell.fill = fill_fo_win
                else: cell.fill = fill_winner

        # Formato de Stock
        for col_st_idx in [col_co_st, col_fa_st, col_fo_st]:
            cell_stock = ws.cell(row=row_idx, column=col_st_idx)
            if cell_stock.value == "Con Stock":
                cell_stock.font = Font(color="047857", bold=True)
            elif cell_stock.value == "Sin Stock":
                cell_stock.font = Font(color="DC2626", bold=True)
            else:
                cell_stock.font = Font(color="9CA3AF") # No encontrado

        # Links
        for label, col_lk_idx in [("CO", col_lk_co), ("FA", col_lk_fa), ("FO", col_lk_fo)]:
            cell_link = ws.cell(row=row_idx, column=col_lk_idx)
            if cell_link.value and str(cell_link.value).startswith("http"):
                cell_link.hyperlink = str(cell_link.value)
                cell_link.value = f"Ver en {label}"
                cell_link.font = font_link

    anchos = [18, 45, 15, 15, 12, 12, 15, 15, 12, 12, 15, 15, 12, 12, 22, 14, 12, 15, 15, 15]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "A2"

print(f"¡Excel comparativo guardado en: {OUTPUT_EXCEL}")
