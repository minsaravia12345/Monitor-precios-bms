import os
import pandas as pd
from glob import glob
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETA_RESULTADOS = os.path.join(BASE_DIR, "Resultados_scraping")
RUTA_KILLERS = r"C:\Users\b_saravia\OneDrive - Farmacity\Documents\skillers\Killers evento.xlsx"
OUTPUT_FALLBACK = os.path.join(BASE_DIR, "killers_urls_fallback.xlsx")

FARMACIAS = [
    ("Central Oeste", "CentralOeste"),
    ("Farmacity",     "Farmacity"),
    ("Farmaonline",   "Farmaonline"),
]

print("--- GENERANDO REPORTE DE PRODUCTOS FALTANTES ---")

if not os.path.exists(RUTA_KILLERS):
    print(f"Error: No se encontró {RUTA_KILLERS}")
    exit()

try:
    df_killers = pd.read_excel(RUTA_KILLERS, dtype={'EAN': str})
    col_nombre = [c for c in df_killers.columns if 'Descripci' in c]
    col_nombre = col_nombre[0] if col_nombre else 'Nombre'
    
    # Limpiamos EANs
    df_killers['EAN_clean'] = df_killers['EAN'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    killers_list = df_killers.to_dict('records')

except Exception as e:
    print(f"Error leyendo {RUTA_KILLERS}: {e}")
    exit()

archivos = glob(os.path.join(CARPETA_RESULTADOS, "*.xlsx"))
archivos = [f for f in archivos if "Comparacion_" not in os.path.basename(f)]

faltantes = []

for nombre_farmacia, patron in FARMACIAS:
    encontrados = sorted([f for f in archivos if patron in f], reverse=True)
    eans_encontrados = set()
    
    if encontrados:
        df_farm = pd.read_excel(encontrados[0])
        if 'EAN' in df_farm.columns:
            # Extraemos todos los EANs que sí se encontraron
            eans_raw = df_farm['EAN'].astype(str).tolist()
            for e in eans_raw:
                if 'e' in e.lower() or '.' in e:
                    try: e = format(float(e), '.0f')
                    except: pass
                e = re.sub(r'\.0$', '', e).strip()
                eans_encontrados.add(e)
    else:
        print(f"Advertencia: No hay resultados previos para {nombre_farmacia}.")

    # Verificamos cuáles Killers faltan en esta farmacia
    count_faltantes = 0
    for k in killers_list:
        ean_k = k['EAN_clean']
        if str(ean_k).lower() == 'nan': ean_k = ''
        
        # Si no tiene EAN, podríamos comparar por nombre, pero para fallback manual es más seguro pedirlo si no hay EAN o si el EAN no está
        if ean_k and ean_k not in eans_encontrados:
            faltantes.append({
                'Farmacia': nombre_farmacia,
                'EAN': ean_k,
                'Nombre': k.get(col_nombre, k.get('Nombre', '')),
                'URL': '' # Columna vacía para que el usuario la llene
            })
            count_faltantes += 1
        elif not ean_k:
            # Si el Killer ni siquiera tiene EAN en el excel original, lo agregamos como faltante para que pongan la URL
            # (a menos que quieran saltarlo, pero mejor prevenir)
            faltantes.append({
                'Farmacia': nombre_farmacia,
                'EAN': '',
                'Nombre': k.get(col_nombre, k.get('Nombre', '')),
                'URL': ''
            })
            count_faltantes += 1
            
    print(f"{nombre_farmacia}: {count_faltantes} productos no encontrados automáticamente.")

if faltantes:
    df_faltantes = pd.DataFrame(faltantes)
    
    # Si ya existe un archivo de fallback previo, preservamos las URLs que ya habían sido ingresadas
    if os.path.exists(OUTPUT_FALLBACK):
        try:
            df_prev = pd.read_excel(OUTPUT_FALLBACK)
            if 'URL' in df_prev.columns and not df_prev.empty:
                # Merge para recuperar las URLs manuales anteriores
                df_prev = df_prev.dropna(subset=['URL'])
                # Hacemos merge por Farmacia y EAN (o Nombre si no hay EAN)
                df_faltantes = pd.merge(df_faltantes, df_prev[['Farmacia', 'EAN', 'Nombre', 'URL']], 
                                        on=['Farmacia', 'EAN', 'Nombre'], 
                                        how='left', suffixes=('', '_prev'))
                # Reemplazamos URLs vacías por las previas
                if 'URL_prev' in df_faltantes.columns:
                    df_faltantes['URL'] = df_faltantes['URL_prev'].combine_first(df_faltantes['URL'])
                    df_faltantes = df_faltantes.drop(columns=['URL_prev'])
        except Exception as e:
            print(f"No se pudieron cargar las URLs previas: {e}")

    # Guardar archivo con formato para que el usuario llene la columna URL
    with pd.ExcelWriter(OUTPUT_FALLBACK, engine="openpyxl") as writer:
        df_faltantes.to_excel(writer, index=False, sheet_name="Faltantes")
        ws = writer.sheets["Faltantes"]
        
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        
        # Pintar cabecera
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            
        # Resaltar la columna URL (4ta)
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 45
        
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                if not cell.value:
                    cell.fill = PatternFill("solid", fgColor="FFFF00") # Amarillo para indicar que falta
                    
    print(f"\nSe generó {OUTPUT_FALLBACK} con {len(df_faltantes)} registros.")
    print("-> Por favor, abre este archivo, pega las URLs correctas en la columna amarilla y vuelve a ejecutar los scrapers.")
else:
    print("\n¡Excelente! Se encontró el 100% de los productos en todas las farmacias.")
    # Si no hay faltantes, podemos limpiar el archivo de fallback o crear uno vacío
    pd.DataFrame(columns=['Farmacia', 'EAN', 'Nombre', 'URL']).to_excel(OUTPUT_FALLBACK, index=False)
